import tempfile, os
from common.storage_azure import get_blob,get_blob_client
from openpyxl import load_workbook
import pyarrow as pa
import pyarrow.parquet as pq
import hashlib, time
from azure.core.exceptions import AzureError
from common.exceptions import HeaderError
import itertools
from logger.logging import setup_logging
logger = setup_logging()
from utility.excel_cleaning import is_valid_header, has_split_headers, is_footer

   
max_attempts = 3
def convert_excel_to_parquet(object_name:str,job_id:str, chunk_size=500) -> str:
    """
    Convert Excel file in S3 → Parquet in S3 (no disk)
    Returns new S3 key
    """
    logger.info(f" [JOB={job_id}]  - Inside funtion - convert excel file {object_name} to parquet and save in Azure Blob")
    writer = None
    excel_path = None
    parquet_path = None
    wb = None
    MAX_HEADER_SCAN = 10 

    try:
 
        stream = get_blob(object_name)

        # Step 1: Download Excel → disk (NO RAM spike)
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_excel:
            stream.readinto(tmp_excel) # streams directly into a tempfile (even better)
            excel_path = tmp_excel.name  # Getting the file path of the temp Excel file

        # Step 3: Write Parquet → disk (NOT memory)
        parquet_path = os.path.splitext(excel_path)[0] + ".parquet"

        # Step 2: Open Excel in streaming mode
        wb = load_workbook(excel_path, read_only=True) # Open Excel in streaming mode. Don’t load the whole Excel — just read it row by row”
        ws = wb.active # This picks the active sheet from the Excel file. “Open the Excel file and go to the currently selected tab”
        rows_iter = ws.iter_rows(values_only=True) # This creates a generator (iterator) that gives rows one at a time.

        #Extract header
        #headers = next(rows_iter, None) # next() is used to get the next item from an iterator and takes that row out of itertaor
        
        #if headers is None:
            #raise ValueError("Excel file is empty")
        
        headers = None
        start_idx = 0
        header_candidates = []

        for _ in range(MAX_HEADER_SCAN):
            row = next(rows_iter, None) 

            if not row:
                continue

            # find first non-empty column
            for i, val in enumerate(row):
                if val is not None and str(val).strip() != "":
                    trimmed = row[i:]
                    start_idx = i
                    break
            else:
                continue

            header_candidates.append(trimmed)

            if is_valid_header(trimmed):
                headers = trimmed
                break
        
     
        if headers is None:
            raise HeaderError("Invalid Excel format: No valid header row detected")
        
        # Read next row
        next_row = next(rows_iter, None)

        if next_row:
            next_row_trimmed = next_row[start_idx:] if start_idx < len(next_row) else []

            if has_split_headers(headers, next_row_trimmed):
                raise HeaderError("Invalid Excel format: Column headers are split across multiple rows")
            

            # Put row back so processing continues normally
            rows_iter = itertools.chain([next_row], rows_iter)


        # Clean headers
        headers = [
                    str(h).strip()
                    for h in headers
                    if h is not None and str(h).strip() != ""
                ]

        batch = [] # Prepare batch storage

        for row in rows_iter:
            try:
                row = row[start_idx:] 
                if is_footer(row):
                    break

                if all(v is None or str(v).strip() == "" for v in row):
                    continue  # skip empty rows
                
                batch.append(row)

            # When chunk is full → write to parquet
                if len(batch) >= chunk_size:
                    table = pa.Table.from_pylist([
                            {
                                h: (str(r[i]) if i < len(r) and r[i] is not None else None)
                                for i, h in enumerate(headers)
                            }
                            for r in batch
                        ])
                    # pa.Table.from_pylist - converts Python data → columnar format (Parquet-ready)

                    if writer is None: 
                        writer = pq.ParquetWriter(parquet_path, table.schema) # Create writer (only once)

                    writer.write_table(table)
                    batch = []

            except Exception:
                logger.exception(f"Error procesisng row {row}")
                continue

        if batch:   
            table = pa.Table.from_pylist([
                            {
                                h: (str(r[i]) if i < len(r) and r[i] is not None else None)
                                for i, h in enumerate(headers)
                            }
                            for r in batch
                        ])

            if writer is None:
                writer = pq.ParquetWriter(parquet_path, table.schema)
            writer.write_table(table)

        if writer:
            writer.close()

        try:
            pf = pq.ParquetFile(parquet_path)
            logger.info(f"Arrow Schema: {pf.schema_arrow}")
            logger.info(f"Schema: {pf.schema}")
        except Exception:
            logger.exception("Parquet validation failed before upload")
            raise
        
        del pf

        # Step 3: Upload Parquet
        
        chunk_hash = hashlib.md5(f"{job_id}".encode()).hexdigest()

        new_key = f"excel_to_parquet/{job_id}/{chunk_hash}.parquet"

        upload_client  = get_blob_client(new_key)

        logger.info(f"Uploading parquet to blob: {new_key}")

        for attempt in range(1, max_attempts + 1):
            try:
                with open(parquet_path, "rb") as f:
                    upload_client.upload_blob(
                        f,
                        overwrite=True,
                        content_type="application/x-parquet"
                    )
                logger.info(f"Upload successful: {new_key}")
                break

            except AzureError:
                if upload_client.exists():
                    logger.info(f"Uploaded parquet to blob: {new_key}")
                    break

                logger.warning(f"Upload failed (attempt {attempt})")

                if attempt == max_attempts:
                    logger.exception("Max upload attempts reached")
                    raise

                time.sleep(2 ** attempt)  

        
        logger.info(f"[JOB={job_id}] - Excel conversion to parquet successful. Uploaded parquet file to blob: {new_key}")

        return new_key
    
    except HeaderError as e:
        logger.exception(f"[JOB={job_id}] {e}")
        raise

    except AzureError:
        logger.exception(f"[JOB={job_id}]  - Parquet file upload to Azure failed")
        raise

    except Exception:
        logger.exception(f"[JOB={job_id}]  - Error in converting...")
        raise
    
    finally:
        if writer:
            try:
                writer.close()
            except:
                pass
 
        if wb:
            wb.close()
       
        # Cleanup temp Excel
        
        if excel_path and os.path.exists(excel_path):
            for attempt in range(3):
                try:
                    os.remove(excel_path)
                    break
                except PermissionError:
                    time.sleep(0.5)

        if parquet_path and os.path.exists(parquet_path):
            os.remove(parquet_path)